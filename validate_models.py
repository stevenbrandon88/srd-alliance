#!/usr/bin/env python3
"""
SRD Alliance — Model Validator
================================
Opens every .xlsx in the repository (or a specified folder) and checks
for formula errors (#REF!, #VALUE!, #NAME?, #DIV/0!, #N/A, #NULL!, #NUM!).

Usage:
    python validate_models.py                     # validates all .xlsx in current dir
    python validate_models.py --dir ./individual  # specific folder
    python validate_models.py --file SRD_Fiji_model.xlsx
    python validate_models.py --strict            # exit code 1 if any errors found

Output:
    Pass/fail table with error counts and locations.
    Summary line with total files, total formulas, total errors.
"""

import os
import sys
import glob
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

ERROR_VALUES = {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!", "#ERROR!"}

# ANSI colours (disabled on Windows if not supported)
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

def supports_colour():
    return sys.stdout.isatty() and os.name != "nt"

def c(colour, text):
    return f"{colour}{text}{RESET}" if supports_colour() else text


def check_workbook(path: Path) -> dict:
    """
    Open a workbook and scan all cells for error values.
    Returns a dict with keys: file, sheets, total_formulas, errors (list of dicts)
    """
    result = {
        "file": path.name,
        "path": str(path),
        "sheets": [],
        "total_formulas": 0,
        "errors": [],
        "load_error": None,
    }

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        result["load_error"] = str(e)
        return result

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                sheet_formulas += 1
                # Check for error string values (data_only=True returns evaluated errors as strings)
                if isinstance(val, str) and val.strip().upper() in ERROR_VALUES:
                    result["errors"].append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "value": val.strip(),
                    })
        result["sheets"].append(sheet_name)
        result["total_formulas"] += sheet_formulas

    return result


def find_xlsx_files(search_dir: Path, recursive: bool = True) -> list:
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    return sorted(search_dir.glob(pattern))


def validate(files: list, strict: bool = False) -> int:
    """Run validation on a list of Path objects. Returns total error count."""

    total_files     = len(files)
    total_formulas  = 0
    total_errors    = 0
    failed_files    = []

    col_w = 52  # file name column width

    print()
    print(c(BOLD, f"SRD Alliance — Model Validator"))
    print(c(CYAN, f"Checking {total_files} file(s)...\n"))
    print(f"  {'File':<{col_w}} {'Formulas':>10}  {'Errors':>8}  {'Status'}")
    print("  " + "─" * (col_w + 32))

    for path in files:
        result = check_workbook(path)

        if result["load_error"]:
            status = c(RED, "LOAD ERROR")
            err_count = "—"
            form_count = "—"
            total_errors += 1
            failed_files.append((path.name, [f"Could not open: {result['load_error']}"]))
        else:
            form_count = result["total_formulas"]
            err_count  = len(result["errors"])
            total_formulas += form_count
            total_errors   += err_count

            if err_count == 0:
                status = c(GREEN, "✓ PASS")
            else:
                status = c(RED, f"✗ FAIL  ({err_count} error{'s' if err_count > 1 else ''})")
                failed_files.append((path.name, result["errors"]))

        print(f"  {path.name:<{col_w}} {str(form_count):>10}  {str(err_count):>8}  {status}")

    print("  " + "─" * (col_w + 32))

    # Summary
    if total_errors == 0:
        summary_status = c(GREEN + BOLD, "ALL PASS")
    else:
        summary_status = c(RED + BOLD, f"{total_errors} ERROR{'S' if total_errors > 1 else ''} FOUND")

    print(f"\n  {c(BOLD, 'TOTAL')}  {total_files} files  |  "
          f"{total_formulas:,} cells checked  |  {summary_status}\n")

    # Detail for failures
    if failed_files:
        print(c(YELLOW, "  Error details:"))
        for fname, errors in failed_files:
            print(f"\n  {c(BOLD, fname)}")
            if isinstance(errors, list) and errors and isinstance(errors[0], str):
                # Load error
                for e in errors:
                    print(f"    {c(RED, '✗')} {e}")
            else:
                for e in errors[:20]:  # cap at 20 per file
                    print(f"    {c(RED, '✗')} Sheet '{e['sheet']}'  Cell {e['cell']}  →  {e['value']}")
                if len(errors) > 20:
                    print(f"    ... and {len(errors) - 20} more errors")
        print()

    if strict and total_errors > 0:
        return 1
    return 0


def main():
    parser = argparse.ArgumentParser(
        description="Validate all SRD Alliance Excel models for formula errors."
    )
    parser.add_argument(
        "--dir", type=Path, default=Path("."),
        help="Directory to search for .xlsx files (default: current directory)"
    )
    parser.add_argument(
        "--file", type=Path, default=None,
        help="Validate a single file instead of a directory"
    )
    parser.add_argument(
        "--no-recurse", action="store_true",
        help="Only check .xlsx in the top-level directory, not subdirectories"
    )
    parser.add_argument(
        "--strict", action="store_true",
        help="Exit with code 1 if any errors are found (useful in CI)"
    )
    args = parser.parse_args()

    if args.file:
        if not args.file.exists():
            print(f"ERROR: File not found: {args.file}")
            sys.exit(1)
        files = [args.file]
    else:
        recursive = not args.no_recurse
        files = find_xlsx_files(args.dir, recursive=recursive)
        if not files:
            print(f"No .xlsx files found in: {args.dir}")
            sys.exit(0)

    exit_code = validate(files, strict=args.strict)
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
